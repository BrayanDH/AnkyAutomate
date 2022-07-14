import random
import time
from pywinauto.keyboard import send_keys
import pyautogui
import keyboard
import string
import random
import openpyxl
import sys

# Pyautogui, simple double click


def DobleClick(LocacionDobleClick1):
    Ejecutado = False
    while Ejecutado == False:
        time.sleep(1)
        Posciciones = pyautogui.locateCenterOnScreen(LocacionDobleClick1)
        if Posciciones != None:
            print("Dando Doble Click")
            pyautogui.doubleClick(Posciciones[0], Posciciones[1])
            time.sleep(1)
            Ejecutado = True
        else:

            print(f"No se encontro la img")

            time.sleep(3)
        LocacionDobleClick = Posciciones
    return LocacionDobleClick


# Pyautogui, simple click
def Click(LocacionClick):
    Ejecutado = False
    while Ejecutado == False:
        time.sleep(1)
        Posciciones = pyautogui.locateCenterOnScreen(LocacionClick, grayscale=True)
        if Posciciones != None:
            print("Dando Click")
            pyautogui.click(Posciciones[0], Posciciones[1])
            time.sleep(1)
            Ejecutado = True
        else:

            print(f"No se encontro la img")

            time.sleep(3)
    LocacionClick = Posciciones
    return LocacionClick

# Pyautogui, Clicks in defined interval of time


def ClickIntervalo(Locacion, Cantidad):
    Ejecutado = False
    while Ejecutado == False:
        time.sleep(1)
        Posciciones = pyautogui.locateCenterOnScreen(Locacion)
        if Posciciones != None:
            print("Dando Click")
            pyautogui.click(Posciciones[0], Posciciones[1])
            pyautogui.click(button='left', clicks=Cantidad, interval=0.1)
            time.sleep(1)
            Ejecutado = True
        else:

            print(f"No se encontro la img")

            time.sleep(3)
    LocacionClick = Posciciones
    return LocacionClick


#Pyautgui, locate in X
def PoscicionadorX(ImagenAbuscar, Movimiento):
    AlmacenaResultadoClick = Click(ImagenAbuscar)
    time.sleep(1)

    Finalizado = pyautogui.dragTo(AlmacenaResultadoClick[0] + Movimiento, AlmacenaResultadoClick[1], 2, button='left')
    return AlmacenaResultadoClick

#Pywinauto, locate in Y


def PoscicionadorY(ImagenAbuscar, Movimiento):
    AlmacenaResultadoClick = Click(ImagenAbuscar)
    time.sleep(1)

    Finalizado = pyautogui.dragTo(AlmacenaResultadoClick[0], AlmacenaResultadoClick[1] + Movimiento, 2, button='left')
    return AlmacenaResultadoClick

# Show counter of time in console


def TimeCounter(Tiempo):
    for Counter in range(Tiempo, 0, -1):
        sys.stdout.write("\r")
        sys.stdout.write("Espera, quedan {:2d} segundos.".format(Counter))
        sys.stdout.flush()
        time.sleep(1)
    if Counter == 1:
        Counter = 0
        sys.stdout.write("\r")
        sys.stdout.write("Espera, quedan {:2d} segundos.\n".
                         format(Counter))
    return Counter

# gen random num in range


def SetNumRandom(Num1, Num2):
    Num = random.randint(Num1, Num2)
    return Num

#Print in funtion


def ShowOnConsole(Show):
    print(Show)
    sys.stdout.write("\r")



# characters to generate password from
characters = list(string.ascii_letters + string.digits + "123456789")


def PassGen():

    length = int(random.randint(9, 12))
    random.shuffle(characters)
    password = []
    for i in range(length):
        password.append(random.choice(characters))
    random.shuffle(password)
    password = "".join(password)

    return password


def PrintACurrentTime():
    now = datetime.now()
    print(now.strftime('%Y/%m/%d %I:%M:%S'))


def GetValue(Book, Row, Column):
    LineC = Row
    Book = Book
    Column = Column
    wb = openpyxl.load_workbook(f'{Book}.xlsx')
    ws = wb.active
    Value = ws.cell(row=LineC, column=Column).value
    wb.close()
    return Value


def VerifyWritesLines(Book, Row, Column):
    LineC = Row
    Book = Book
    Column = Column
    VerifyUse = True
    Use = "Yes"
    wb = openpyxl.load_workbook(f'{Book}.xlsx')
    ws = wb.active
    while VerifyUse == True:
        ProbeUse = ws.cell(row=LineC, column=Column).value
        if ProbeUse == "Yes":
            LineC += 1
        elif ProbeUse != "Yes":
            Data = ws.cell(row=LineC, column=Column).value
            VerifyUse = False
    wb.close()
    return Data
